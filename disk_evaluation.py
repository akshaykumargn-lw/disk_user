#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Install required libraries from 'requirements' file:
import sys
import os
import openpyxl
import pandas as pd
import datetime
from pathlib import Path

extensions_list = []
selected_folder = ''
size_limit = []
size_limit_arg = []

def main():
    global extensions_list, selected_folder, size_limit, size_limit_arg  # Declare these variables as global
    if len(sys.argv) != 4:
        print("Usage: program.py <extensions.txt> </CAE folder> <file size>")
        return

    extensions_file = sys.argv[1]
    selected_folder = sys.argv[2]
    size_limit = sys.argv[3]

    # Check if the provided file and folder paths exist
    if not os.path.isfile(extensions_file):
        print(f"The extensions file '{extensions_file}' does not exist.")
        return

    if not os.path.isdir(selected_folder):
        print(f"The folder '{selected_folder}' does not exist.")
        return

    print(f"Extensions file: {extensions_file}")
    print(f"selected folder: {selected_folder}")
    
    try:
        # Try to parse the argument with units (KB, MB, GB)
        size_limit_arg, unit = size_limit[:-2], size_limit[-2:].upper()
        size_limit_arg = int(float(size_limit_arg) * {'KB': 1024, 'MB': 1024 ** 2, 'GB': 1024 ** 3}[unit])
        print(f"Size limit: {sys.argv[3]}")
    except ValueError:
        print("Invalid file size format. Use 'KB', 'MB', 'GB'.")

    # Read file extensions from the extension list file
    try:
        with open(extensions_file, 'r') as f:
            extensions_list = [line.strip() for line in f]
            if all(ext.startswith("*.") for ext in extensions_list):
                print("Extensions read from file:", extensions_list)          
            else:
                print("Invalid input. Each line in the file should start with '*.ext'.")
    except FileNotFoundError:
        print("File not found. Please enter a valid file name.")
    
if __name__ == "__main__":
    main()    
print("Process started. Please wait!!")
# Define a dictionary to map unit abbreviations to multipliers (e.g., KB to 1024)
unit_multipliers = {
    'B': 1,
    'KB': 1024,
    'MB': 1024 ** 2,
    'GB': 1024 ** 3,
    'TB': 1024 ** 4,
}
        
# Convert the list to a comma-separated string
file_ext_str = ",".join(extensions_list)

# Create a list to store the matching file paths
matching_files = []

# Use pathlib to search for matching files
folder_path = Path(selected_folder)

for extension in extensions_list:
    matching_files.extend(folder_path.glob('**/' + extension))

# Function to get file author
def get_file_author(file_path):
    try:
        # Use a platform-specific method to get file author (example for Linux)
        author = os.popen(f"stat -c '%U' '{file_path}'").read().strip()
        return author
    except:
        return "N/A"

# Function to get file size
def get_file_size(file_path):
    try:
        # Check if the file exists
        if not os.path.exists(file_path):
           return f'File not found- {file_path}' # Zero file size is assigned for invalid file types
        size = os.path.getsize(file_path)
        return size
    except Exception as e:
        print(f"Error getting file size: {e}")
        return "N/A"

#Function to get file latest modified time stamp
def get_last_modified_timestamp(file_path):
    try:
        # Get the last modified timestamp
        timestamp = os.path.getmtime(file_path)
        # Convert the timestamp to a datetime object
        modified_date = datetime.datetime.fromtimestamp(timestamp)
       # Format the datetime as a string in the desired format (e.g., MM/DD/YYYY HH:MM:SS)
        formatted_date = modified_date.strftime("%m/%d/%Y %H:%M")
        if formatted_date.endswith('.0'):
            formatted_date = formatted_date[:-2]
    except FileNotFoundError:
        # Handle the FileNotFoundError by setting a "NAN" value
        modified_date = "NAN"
        
    return modified_date
      
# Create a list to store the file information
file_info = []

# Extract and store file information in the file_info list
for file_path in matching_files:
    file_name = file_path.name
    file_author = get_file_author(file_path)
    file_size = get_file_size(file_path)
    last_modified_timestamp = get_last_modified_timestamp(file_path
                                                          )
    # Encode the file path to UTF-8 to handle non-UTF-8 characters
    utf8_file_path = str(file_path).encode('utf-8', 'ignore').decode('utf-8')
    
    file_info.append((file_name, file_author, file_size, utf8_file_path, last_modified_timestamp))
 
    # Create a Pandas DataFrame from the file_info list
df0 = pd.DataFrame(file_info, columns=["File Name", "Author", "Size", "File Path", "Modified Timestamp"])
#Defining timestamp column as datetime64[ns] datatype
df0["Modified Timestamp"] = pd.to_datetime(df0["Modified Timestamp"], format="%Y-%m-%d %H:%M:%S")

# Save the Excel file with both selected file type and selected folder in the filename
file_ext_str = file_ext_str.replace("*", "")  # Remove asterisk from the file type
 
print(f"Found {len(file_info)} matching files.")

# Function to clean and convert the string in 'Size' column
def process_size_column(df):
    for index, row in df0.iterrows():
        size_value = row['Size']
        if isinstance(size_value, str):
            print(f"Found a string value in row {index}: {size_value}")
            df0.at[index, 'Size'] = 0
    return df0
       
df1 = process_size_column(df0)

df2 = df1[df1['Size'] >= size_limit_arg]  # Filter based on the numeric 'Size'

# Calculate the total size in the SIZE column
total_size_byte = df2['Size'].sum()

def convert_size(dataframe):
    for index, row in dataframe.iterrows():
        size_value = row['Size']
        if isinstance(size_value, (int, float)):
            if size_value >= 1e9:
                dataframe.at[index, 'Size'] = f"{size_value / 1e9:.2f}GB"
            elif size_value >= 1e6:
                dataframe.at[index, 'Size'] = f"{size_value / 1e6:.2f}MB"
            elif size_value >= 1e3:
                dataframe.at[index, 'Size'] = f"{size_value / 1e3:.2f}KB"
            else:
                dataframe.at[index, 'Size'] = f"{size_value:.2f}bytes"
    return dataframe

#Sort the DataFrame by file size in descending order
df2 = df2.sort_values(by='Size', ascending=False)

df3 = df2.copy()

# Convert the 'Size' column
df3 = convert_size(df3)

# Function to replace non-UTF-8 characters with '?'
def replace_non_utf8_characters(text):
    if isinstance(text, str):
        try:
            text.encode('utf-8', errors='surrogatepass').decode('utf-8')
            return text  # Text is already valid UTF-8
        except UnicodeDecodeError:
            return '?'  # Replace non-UTF-8 characters with '?'
    else:
        return text  # For Timestamp
    
# Create a new DataFrame with replaced characters
df4 = df3.applymap(replace_non_utf8_characters)

# Function to format size in a human-readable way (KB, MB, or GB)
def format_size(size_in_bytes):
    # Define size units
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    unit_index = 0

    # Convert to the appropriate unit (KB, MB, GB)
    while size_in_bytes >= 1024 and unit_index < len(units) - 1:
        size_in_bytes /= 1024
        unit_index += 1

    return f'Size: {size_in_bytes:.2f} {units[unit_index]}'

# Format total_size in a human-readable way
total_size_formatted = format_size(total_size_byte)

# Update the column header with the total size
df4 = df4.rename(columns={'Size': total_size_formatted})

# Get the current date and time
current_date = datetime.datetime.now()

# Format the date as a string in the desired format (e.g., YYYY-MM-DD)
date_string = current_date.strftime("%Y-%m-%d")

# Sanitize the selected_folder variable
sanitized_selected_folder = selected_folder.replace('/', '_')

# Create a Pandas Excel writer object for saving the updated data
updated_excel_filename = f'{date_string},Extensions={file_ext_str},Search_Folders={sanitized_selected_folder},Filter_Size={size_limit}.xlsx'

with pd.ExcelWriter(updated_excel_filename, engine='openpyxl') as writer:
    df4.to_excel(writer, sheet_name='All Users', index=False)
   
    # Group the data by author name and save each group to a separate sheet
    grouped = df4.groupby('Author')

    for author, group_data in grouped:
        sheet_name = f'{author}'
        group_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
    # Get the openpyxl workbook object and save it
    workbook = writer.book
    workbook.save(updated_excel_filename)
    
df6 = df2.applymap(replace_non_utf8_characters)
# Create a dictionary to store the sum for each author
author_sums = {}
formatted_sums = {}  # Create a dictionary to store the formatted sums

grouped_data = df6.groupby('Author')
for author, group_data in grouped_data:
    author_sums[author] = group_data['Size'].sum()
    formatted_sums[author] = format_size(author_sums[author])
    
# Load the Excel file to apply the column heading changes
    wb = openpyxl.load_workbook(updated_excel_filename)

    # Access and modify the column heading cell in each sheet (except the first sheet)
    for sheet_name in wb.sheetnames[1:]:
        ws = wb[sheet_name]
        author_name = sheet_name  
        if author_name in formatted_sums:
            size_sum = formatted_sums[author_name]
            ws.cell(row=1, column=3, value=size_sum)  #  Modifying column C (C1)

    # Save the modified Excel file
    wb.save(updated_excel_filename)
    
print(f"Data has been segregated by author and saved to '{updated_excel_filename}'.")
print("Process complete")